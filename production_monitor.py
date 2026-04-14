#!/usr/bin/env python3
"""
Production Multi-Agent Research Monitoring System
No LLM - Pure Automation with Advanced Filtering
Ready for KVM-1 VPS Deployment
"""

import requests
import time
import sqlite3
from typing import List, Dict, Optional, TypedDict, Annotated
from datetime import datetime, timedelta
from pathlib import Path
import operator
import json
import urllib.request
import xml.etree.ElementTree as ET
import hashlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from langgraph.graph import StateGraph, END
from apscheduler.schedulers.blocking import BlockingScheduler

import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================================
# CONFIGURATION
# ================================
class Config:
    # Search Configuration
    BASE_QUERY = "computer vision"
    LIMIT = 20
    
    # Source Selection
    ENABLE_SEMANTIC_SCHOLAR = True
    ENABLE_ARXIV = True  # Set to False to disable arXiv
    
    # API Settings
    SEMANTIC_SCHOLAR_URL = "https://api.semanticscholar.org/graph/v1"
    ARXIV_API_URL = "http://export.arxiv.org/api/query"
    TIMEOUT = 10
    RATE_LIMIT = 1  # Seconds between API calls
    
    # Venue Rankings
    VENUE_RANKS = {
        "CVPR": "A*", "ICCV": "A*", "ECCV": "A*",
        "NeurIPS": "A*", "ICML": "A*", "ICLR": "A*",
        "AAAI": "A", "IJCAI": "A",
        "TPAMI": "Q1", "IJCV": "Q1", "TIP": "Q1",
        "BMVC": "B", "WACV": "B"
    }
    
    # Storage
    DB_PATH = "research_papers.db"
    FILTERS_PATH = "user_filters.json"
    REPORT_DIR = Path("reports")
    
    # Telegram Bot (create via @BotFather). Bots cannot send by phone number — only chat_id.
    # Get chat_id: open the bot, send /start, then GET .../getUpdates and read message.chat.id
    TELEGRAM_BOT_TOKEN = "000000000:XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
    TELEGRAM_CHAT_ID = "000000000"
    
    # Schedule
    RUN_HOUR = 9
    RUN_MINUTE = 0

# ================================
# USER FILTERS
# ================================
class UserFilters:
    """Manage user-defined filters for paper selection"""
    
    @staticmethod
    def load():
        """Load filters from JSON file"""
        if Path(Config.FILTERS_PATH).exists():
            with open(Config.FILTERS_PATH, 'r') as f:
                return json.load(f)
        
        # Default filters (all disabled)
        return {
            "enabled": True,
            "query": Config.BASE_QUERY,
            
            # Author filters
            "author_names": [],  # e.g., ["Yann LeCun", "Fei-Fei Li"]
            "author_institutions": [],  # e.g., ["MIT", "Stanford", "Google"]
            
            # Venue filters
            "venues": [],  # e.g., ["CVPR", "ICCV", "NeurIPS"]
            "min_venue_rank": None,  # "A*", "Q1", "A", "B", or None
            
            # Citation filters
            "min_citations": 0,
            "max_citations": None,
            
            # Date filters
            "min_year": None,
            "max_year": None,
            "last_n_days": None,  # Papers from last N days
            
            # Content filters
            "keywords_include": [],  # Must contain at least one
            "keywords_exclude": [],  # Must not contain any
            
            # Quantity
            "max_results": 20
        }
    
    @staticmethod
    def save(filters: Dict):
        """Save filters to file"""
        with open(Config.FILTERS_PATH, 'w') as f:
            json.dump(filters, f, indent=2)
        logger.info(f"Filters saved to {Config.FILTERS_PATH}")
    
    @staticmethod
    def apply_to_papers(papers: List[Dict], filters: Dict) -> List[Dict]:
        """Apply all filters to paper list"""
        if not filters.get("enabled", True):
            return papers
        
        filtered = papers
        
        # Author name filter
        if filters.get("author_names"):
            author_names_lower = [a.lower() for a in filters["author_names"]]
            filtered = [
                p for p in filtered
                if any(
                    any(filter_name in author.lower() for filter_name in author_names_lower)
                    for author in p.get("authors", [])
                )
            ]
        
        # Institution filter
        if filters.get("author_institutions"):
            institutions_lower = [i.lower() for i in filters["author_institutions"]]
            filtered = [
                p for p in filtered
                if any(
                    any(inst in detail.get("affiliation", "").lower() for inst in institutions_lower)
                    for detail in p.get("author_details", [])
                )
            ]
        
        # Venue filter
        if filters.get("venues"):
            venues_lower = [v.lower() for v in filters["venues"]]
            filtered = [
                p for p in filtered
                if any(venue in p.get("venue", "").lower() for venue in venues_lower)
            ]
        
        # Venue rank filter
        if filters.get("min_venue_rank"):
            rank_order = {"A*": 4, "Q1": 3, "A": 2, "B": 1, "Unranked": 0}
            min_rank_value = rank_order.get(filters["min_venue_rank"], 0)
            filtered = [
                p for p in filtered
                if rank_order.get(p.get("venue_rank", "Unranked"), 0) >= min_rank_value
            ]
        
        # Citation filters
        if filters.get("min_citations") is not None:
            filtered = [p for p in filtered if p.get("citations", 0) >= filters["min_citations"]]
        
        if filters.get("max_citations") is not None:
            filtered = [p for p in filtered if p.get("citations", 0) <= filters["max_citations"]]
        
        # Year filters
        if filters.get("min_year"):
            filtered = [p for p in filtered if p.get("year", 0) >= filters["min_year"]]
        
        if filters.get("max_year"):
            filtered = [p for p in filtered if p.get("year", 9999) <= filters["max_year"]]
        
        # Last N days filter
        if filters.get("last_n_days"):
            cutoff_date = datetime.now() - timedelta(days=filters["last_n_days"])
            cutoff_str = cutoff_date.strftime("%Y-%m-%d")
            filtered = [
                p for p in filtered
                if p.get("publication_date", "9999-12-31") >= cutoff_str
            ]
        
        # Keyword inclusion
        if filters.get("keywords_include"):
            keywords_lower = [k.lower() for k in filters["keywords_include"]]
            filtered = [
                p for p in filtered
                if any(
                    keyword in p.get("title", "").lower() or 
                    keyword in p.get("abstract", "").lower()
                    for keyword in keywords_lower
                )
            ]
        
        # Keyword exclusion
        if filters.get("keywords_exclude"):
            keywords_lower = [k.lower() for k in filters["keywords_exclude"]]
            filtered = [
                p for p in filtered
                if not any(
                    keyword in p.get("title", "").lower() or 
                    keyword in p.get("abstract", "").lower()
                    for keyword in keywords_lower
                )
            ]
        
        # Max results
        if filters.get("max_results"):
            filtered = filtered[:filters["max_results"]]
        
        return filtered

# ================================
# STATE DEFINITION
# ================================
class ResearchState(TypedDict):
    filters: Dict
    papers: Annotated[List[Dict], operator.add]
    semantic_scholar_papers: List[Dict]  # From Semantic Scholar
    arxiv_papers: List[Dict]  # From arXiv
    new_papers: List[Dict]
    dedupe_new_count: int  # papers whose IDs were not in DB (before fallback batch)
    filtered_papers: List[Dict]
    report_path: Optional[str]
    email_sent: bool
    error: Optional[str]
    stats: Dict
    timestamp: str

# ================================
# AGENT NODES
# ================================

def fetch_arxiv_papers_node(state: ResearchState) -> ResearchState:
    """
    Agent 1a: Fetch papers from arXiv
    """
    if not Config.ENABLE_ARXIV:
        logger.info("⏭️  arXiv disabled")
        return {"arxiv_papers": []}
    
    logger.info("📚 Agent 1a: Fetching papers from arXiv...")
    
    filters = state["filters"]
    query = filters.get("query", Config.BASE_QUERY)
    
    # arXiv search categories for computer vision
    # cat:cs.CV = Computer Vision, cat:cs.AI = AI, cat:cs.LG = Machine Learning
    search_query = f"all:{query}"
    
    # Add category filters based on query
    if "computer vision" in query.lower() or "cv" in query.lower():
        search_query = f"cat:cs.CV AND ({query})"
    elif "machine learning" in query.lower() or "deep learning" in query.lower():
        search_query = f"(cat:cs.LG OR cat:cs.CV) AND ({query})"
    
    # Build URL
    url = f"{Config.ARXIV_API_URL}?search_query={urllib.parse.quote(search_query)}"
    url += f"&max_results={Config.LIMIT}"
    url += "&sortBy=submittedDate&sortOrder=descending"
    
    try:
        with urllib.request.urlopen(url, timeout=Config.TIMEOUT) as response:
            xml_data = response.read()
        
        # Parse XML
        root = ET.fromstring(xml_data)
        ns = {'atom': 'http://www.w3.org/2005/Atom', 'arxiv': 'http://arxiv.org/schemas/atom'}
        
        structured = []
        
        for entry in root.findall('atom:entry', ns):
            # Extract data
            title = entry.find('atom:title', ns).text.strip().replace('\n', ' ')
            
            # Authors
            authors = [
                author.find('atom:name', ns).text 
                for author in entry.findall('atom:author', ns)
            ]
            
            # arXiv ID
            arxiv_id = entry.find('atom:id', ns).text.split('/abs/')[-1]
            
            # Published date
            published = entry.find('atom:published', ns).text[:10]  # YYYY-MM-DD
            
            # Abstract
            abstract = entry.find('atom:summary', ns).text.strip().replace('\n', ' ')
            
            # PDF link
            pdf_link = f"https://arxiv.org/pdf/{arxiv_id}.pdf"
            
            # Categories
            categories = [
                cat.attrib['term'] 
                for cat in entry.findall('atom:category', ns)
            ]
            primary_category = entry.find('arxiv:primary_category', ns)
            primary_cat = primary_category.attrib['term'] if primary_category is not None else categories[0] if categories else "Unknown"
            
            # Create paper ID (hash of arXiv ID for consistency)
            paper_id = f"arxiv_{arxiv_id.replace('.', '_')}"
            
            # Parse year from published date
            year = int(published.split('-')[0]) if published else None
            
            structured.append({
                "paper_id": paper_id,
                "arxiv_id": arxiv_id,
                "title": title,
                "authors": authors,
                "author_ids": [],  # arXiv doesn't provide author IDs
                "venue": f"arXiv:{primary_cat}",
                "year": year,
                "publication_date": published,
                "citations": 0,  # arXiv doesn't track citations
                "abstract": abstract[:500],
                "pdf_link": pdf_link,
                "author_details": [],
                "venue_rank": "Preprint",  # arXiv papers are preprints
                "is_new": True,
                "source": "arxiv"
            })
        
        logger.info(f"✅ Fetched {len(structured)} papers from arXiv")
        return {"arxiv_papers": structured, "papers": structured}
        
    except Exception as e:
        logger.error(f"❌ arXiv fetch failed: {e}")
        return {"arxiv_papers": [], "error": str(e)}


def fetch_papers_node(state: ResearchState) -> ResearchState:
    """
    Agent 1b: Fetch papers from Semantic Scholar
    """
    if not Config.ENABLE_SEMANTIC_SCHOLAR:
        logger.info("⏭️  Semantic Scholar disabled")
        return {"semantic_scholar_papers": []}
    
    logger.info("🔍 Agent 1b: Fetching papers from Semantic Scholar...")
    
    filters = state["filters"]
    query = filters.get("query", Config.BASE_QUERY)
    
    url = f"{Config.SEMANTIC_SCHOLAR_URL}/paper/search"
    
    # Build query with filters
    search_query = query
    
    # Add venue filter to query if specified
    if filters.get("venues") and len(filters["venues"]) <= 3:
        venue_filter = " OR ".join([f'venue:"{v}"' for v in filters["venues"]])
        search_query = f"{query} ({venue_filter})"
    
    params = {
        "query": search_query,
        "limit": min(filters.get("max_results", 20) * 2, 100),  # Fetch extra for filtering
        "fields": "paperId,title,authors,year,venue,citationCount,publicationDate,abstract",
        "year": f"{filters.get('min_year', 2000)}-{filters.get('max_year', 2030)}" if filters.get("min_year") or filters.get("max_year") else None
    }
    
    # Remove None values
    params = {k: v for k, v in params.items() if v is not None}
    
    try:
        response = requests.get(url, params=params, timeout=Config.TIMEOUT)
        response.raise_for_status()
        
        raw_papers = response.json().get("data", [])
        
        # Structure papers
        structured = []
        for paper in raw_papers:
            authors = []
            author_ids = []
            
            for a in paper.get("authors", []):
                authors.append(a.get("name", "Unknown"))
                if a.get("authorId"):
                    author_ids.append(a["authorId"])
            
            structured.append({
                "paper_id": paper.get("paperId"),
                "title": paper.get("title", "Unknown"),
                "authors": authors,
                "author_ids": author_ids,
                "venue": paper.get("venue"),
                "year": paper.get("year"),
                "publication_date": paper.get("publicationDate"),
                "citations": paper.get("citationCount", 0),
                "abstract": paper.get("abstract", "")[:500],  # Truncate abstract
                "author_details": [],
                "venue_rank": "Unknown",
                "is_new": True
            })
        
        logger.info(f"✅ Fetched {len(structured)} papers")
        return {"semantic_scholar_papers": structured, "papers": structured, "error": None}
        
    except Exception as e:
        logger.error(f"❌ Fetch failed: {e}")
        return {"semantic_scholar_papers": [], "error": str(e)}


def filter_duplicates_node(state: ResearchState) -> ResearchState:
    """
    Agent 2: Merge papers from both sources and filter duplicates
    """
    logger.info("🔍 Agent 2: Merging sources and filtering duplicates...")
    
    # Combine papers from both sources
    all_papers = []
    all_papers.extend(state.get("semantic_scholar_papers", []))
    all_papers.extend(state.get("arxiv_papers", []))
    
    logger.info(f"Combined {len(all_papers)} papers from all sources")
    
    conn = sqlite3.connect(Config.DB_PATH)
    cursor = conn.cursor()
    
    # Create tables if not exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS papers (
            paper_id TEXT PRIMARY KEY,
            title TEXT,
            venue TEXT,
            year INTEGER,
            citations INTEGER,
            venue_rank TEXT,
            publication_date TEXT,
            discovered_date TEXT
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS authors (
            author_name TEXT,
            paper_id TEXT,
            h_index INTEGER,
            paper_count INTEGER,
            total_citations INTEGER,
            affiliation TEXT,
            FOREIGN KEY(paper_id) REFERENCES papers(paper_id)
        )
    """)
    
    cursor.execute("SELECT paper_id FROM papers")
    existing_ids = {row[0] for row in cursor.fetchall()}
    conn.close()
    
    # Filter new papers (deduplication unchanged)
    new_papers = [p for p in all_papers if p["paper_id"] not in existing_ids]
    dedupe_new_count = len(new_papers)

    # Fallback: no new IDs vs DB but we have a fetch batch — continue pipeline for report email
    if not new_papers and all_papers:
        new_papers = list(all_papers)
        logger.info(
            "No new paper IDs vs database; continuing with fetched papers for report pipeline (fallback)"
        )

    logger.info(f"✅ Found {dedupe_new_count} new papers (out of {len(all_papers)} total)")
    return {"new_papers": new_papers, "papers": all_papers, "dedupe_new_count": dedupe_new_count}


def enrich_authors_node(state: ResearchState) -> ResearchState:
    """
    Agent 3: Enrich author details (only for papers that pass initial filters)
    """
    logger.info("👥 Agent 3: Enriching author profiles...")
    
    papers = state["new_papers"]
    filters = state["filters"]
    
    # Skip if no institution filter (save API calls)
    if not filters.get("author_institutions"):
        logger.info("⏭️  Skipping author enrichment (no institution filter)")
        return {"new_papers": papers}
    
    enriched_count = 0
    
    for paper in papers:
        # Only enrich first 3 authors to save API calls
        for author_id in paper.get("author_ids", [])[:3]:
            time.sleep(Config.RATE_LIMIT)
            
            url = f"{Config.SEMANTIC_SCHOLAR_URL}/author/{author_id}"
            params = {"fields": "name,hIndex,paperCount,citationCount,affiliations"}
            
            try:
                response = requests.get(url, params=params, timeout=Config.TIMEOUT)
                response.raise_for_status()
                
                data = response.json()
                affiliations = data.get("affiliations", [])
                
                paper["author_details"].append({
                    "name": data.get("name", "Unknown"),
                    "h_index": data.get("hIndex", 0),
                    "paper_count": data.get("paperCount", 0),
                    "citations": data.get("citationCount", 0),
                    "affiliation": affiliations[0] if affiliations else "Unknown"
                })
                enriched_count += 1
                
            except Exception as e:
                logger.warning(f"⚠️  Failed to enrich author {author_id}: {e}")
    
    logger.info(f"Enriched {enriched_count} author profiles")
    return {"new_papers": papers}


def apply_filters_node(state: ResearchState) -> ResearchState:
    """
    Agent 4: Apply user-defined filters
    """
    logger.info("Agent 4: Applying user filters...")
    
    papers = state["new_papers"]
    filters = state["filters"]
    
    filtered = UserFilters.apply_to_papers(papers, filters)
    
    logger.info(f"{len(filtered)} papers passed filters (from {len(papers)})")
    
    # Calculate stats (new_papers = count actually new vs DB, not fallback batch size)
    stats = {
        "total_fetched": len(state.get("papers", [])),
        "new_papers": state.get("dedupe_new_count", len(papers)),
        "after_filters": len(filtered),
        "filtered_out": len(papers) - len(filtered)
    }
    
    return {"filtered_papers": filtered, "stats": stats}


def analyze_impact_node(state: ResearchState) -> ResearchState:
    """
    Agent 5: Analyze venue impact factors
    """
    logger.info("Agent 5: Analyzing venue impact...")
    
    papers = state["filtered_papers"]
    
    for paper in papers:
        # Skip if already marked as Preprint (arXiv)
        if paper.get("venue_rank") == "Preprint":
            continue
        
        venue = paper.get("venue", "")
        rank = "Unranked"
        
        if venue:
            venue_lower = venue.lower()
            for conf, conf_rank in Config.VENUE_RANKS.items():
                if conf.lower() in venue_lower:
                    rank = conf_rank
                    break
        
        paper["venue_rank"] = rank
    
    logger.info("Impact analysis complete")
    return {"filtered_papers": papers}


def store_data_node(state: ResearchState) -> ResearchState:
    """
    Agent 6: Store papers in database
    """
    logger.info("Agent 6: Storing papers in database...")
    
    papers = state["filtered_papers"]
    
    if not papers:
        logger.info("No papers to store")
        return {}
    
    conn = sqlite3.connect(Config.DB_PATH)
    cursor = conn.cursor()
    
    discovered_date = datetime.now().strftime("%Y-%m-%d")
    stored_count = 0
    
    for paper in papers:
        try:
            # Insert paper
            cursor.execute("""
                INSERT OR IGNORE INTO papers 
                (paper_id, title, venue, year, citations, venue_rank, publication_date, discovered_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                paper["paper_id"],
                paper["title"],
                paper.get("venue"),
                paper.get("year"),
                paper["citations"],
                paper["venue_rank"],
                paper.get("publication_date"),
                discovered_date
            ))
            
            # Insert authors
            for author in paper.get("author_details", []):
                cursor.execute("""
                    INSERT INTO authors 
                    (author_name, paper_id, h_index, paper_count, total_citations, affiliation)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    author["name"],
                    paper["paper_id"],
                    author["h_index"],
                    author["paper_count"],
                    author["citations"],
                    author["affiliation"]
                ))
            
            stored_count += 1
            
        except sqlite3.Error as e:
            logger.error(f"Failed to store {paper['paper_id']}: {e}")
    
    conn.commit()
    conn.close()
    
    logger.info(f"Stored {stored_count} papers")
    return {}


def generate_excel_node(state: ResearchState) -> ResearchState:
    """
    Agent 7: Generate Excel report
    """
    logger.info("Agent 7: Generating Excel report...")
    
    papers = state["filtered_papers"]
    stats = state.get("stats", {})
    filters = state["filters"]
    
    if not papers:
        logger.info("No papers to report")
        return {"report_path": None}
    
    Config.REPORT_DIR.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = Config.REPORT_DIR / f"research_report_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # ==================
    # SHEET 1: Summary
    # ==================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = "Research Monitoring Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary.merge_cells('A1:B1')
    
    # Stats
    ws_summary.append([])
    ws_summary.append(["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_summary.append(["Query:", filters.get("query", "N/A")])
    ws_summary.append([])
    ws_summary.append(["Statistics", ""])
    ws_summary.append(["Total Papers Fetched:", stats.get("total_fetched", 0)])
    ws_summary.append(["New Papers:", stats.get("new_papers", 0)])
    ws_summary.append(["After Filters:", stats.get("after_filters", 0)])
    ws_summary.append(["Filtered Out:", stats.get("filtered_out", 0)])
    
    # Active filters
    ws_summary.append([])
    ws_summary.append(["Active Filters", ""])
    
    if filters.get("author_names"):
        ws_summary.append(["Authors:", ", ".join(filters["author_names"])])
    if filters.get("author_institutions"):
        ws_summary.append(["Institutions:", ", ".join(filters["author_institutions"])])
    if filters.get("venues"):
        ws_summary.append(["Venues:", ", ".join(filters["venues"])])
    if filters.get("min_venue_rank"):
        ws_summary.append(["Min Venue Rank:", filters["min_venue_rank"]])
    if filters.get("min_citations"):
        ws_summary.append(["Min Citations:", filters["min_citations"]])
    if filters.get("min_year") or filters.get("max_year"):
        ws_summary.append(["Year Range:", f"{filters.get('min_year', 'N/A')} - {filters.get('max_year', 'N/A')}"])
    
    # Venue breakdown
    ws_summary.append([])
    ws_summary.append(["Venue Rank Distribution", "Count"])
    
    rank_counts = {}
    for paper in papers:
        rank = paper["venue_rank"]
        rank_counts[rank] = rank_counts.get(rank, 0) + 1
    
    for rank in ["A*", "Q1", "A", "B", "Unranked"]:
        if rank in rank_counts:
            ws_summary.append([rank, rank_counts[rank]])
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40
    
    # ==================
    # SHEET 2: Papers
    # ==================
    ws_papers = wb.create_sheet("Papers")
    
    headers = ["#", "Title", "Authors", "Venue", "Rank", "Year", "Citations", "Pub Date"]
    ws_papers.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(headers) + 1):
        cell = ws_papers.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Add papers
    for i, paper in enumerate(papers, 1):
        authors_str = ", ".join(paper.get("authors", [])[:3])
        if len(paper.get("authors", [])) > 3:
            authors_str += " et al."
        
        # Add arXiv link if available
        venue_display = paper.get("venue", "N/A")
        if paper.get("source") == "arxiv" and paper.get("arxiv_id"):
            venue_display = f"arXiv ({paper['arxiv_id']})"
        
        row = [
            i,
            paper["title"],
            authors_str,
            venue_display,
            paper["venue_rank"],
            paper.get("year", "N/A"),
            paper["citations"],
            paper.get("publication_date", "N/A")
        ]
        ws_papers.append(row)
        
        # Apply borders
        for col in range(1, len(row) + 1):
            ws_papers.cell(i + 1, col).border = thin_border
    
    # Adjust widths
    ws_papers.column_dimensions['A'].width = 5
    ws_papers.column_dimensions['B'].width = 60
    ws_papers.column_dimensions['C'].width = 30
    ws_papers.column_dimensions['D'].width = 25
    ws_papers.column_dimensions['E'].width = 10
    ws_papers.column_dimensions['F'].width = 8
    ws_papers.column_dimensions['G'].width = 12
    ws_papers.column_dimensions['H'].width = 12
    
    # ==================
    # SHEET 3: Authors (if enriched)
    # ==================
    if any(p.get("author_details") for p in papers):
        ws_authors = wb.create_sheet("Author Details")
        
        author_headers = ["Paper", "Author", "Institution", "h-index", "Papers", "Citations"]
        ws_authors.append(author_headers)
        
        for col in range(1, len(author_headers) + 1):
            cell = ws_authors.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for paper in papers:
            for author in paper.get("author_details", []):
                ws_authors.append([
                    paper["title"][:40] + "...",
                    author["name"],
                    author["affiliation"],
                    author["h_index"],
                    author["paper_count"],
                    author["citations"]
                ])
        
        ws_authors.column_dimensions['A'].width = 45
        ws_authors.column_dimensions['B'].width = 25
        ws_authors.column_dimensions['C'].width = 35
        ws_authors.column_dimensions['D'].width = 10
        ws_authors.column_dimensions['E'].width = 10
        ws_authors.column_dimensions['F'].width = 12
    
    # Save
    wb.save(report_path)
    
    logger.info(f"Report saved: {report_path}")
    return {"report_path": str(report_path)}


def _telegram_send_document(data: bytes, filename: str, caption: Optional[str] = None) -> None:
    url = f"https://api.telegram.org/bot{Config.TELEGRAM_BOT_TOKEN}/sendDocument"
    files = {"document": (filename, data)}
    form: Dict = {"chat_id": Config.TELEGRAM_CHAT_ID}
    if caption:
        form["caption"] = caption[:1024]
    r = requests.post(url, data=form, files=files, timeout=max(Config.TIMEOUT, 120))
    r.raise_for_status()


def send_email_node(state: ResearchState) -> ResearchState:
    """
    Agent 8: Send email notification
    """
    logger.info("Agent 8: Sending Telegram notification...")
    
    papers = state["filtered_papers"]
    report_path = state.get("report_path")
    stats = state.get("stats", {})
    filters = state["filters"]
    
    if not papers:
        logger.info("No papers - skipping notification")
        return {"email_sent": False}
    
    try:
        # Rank breakdown
        rank_counts = {}
        for paper in papers:
            rank = paper["venue_rank"]
            rank_counts[rank] = rank_counts.get(rank, 0) + 1
        
        # Build HTML
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #4472C4; color: white; padding: 15px; }}
                .stats {{ background-color: #f5f5f5; padding: 10px; margin: 10px 0; }}
                .paper {{ margin: 15px 0; padding: 10px; border-left: 3px solid #4472C4; }}
                .badge {{ 
                    display: inline-block; 
                    padding: 3px 8px; 
                    border-radius: 3px; 
                    font-size: 11px; 
                    font-weight: bold;
                }}
                .rank-as {{ background-color: #28a745; color: white; }}
                .rank-q1 {{ background-color: #17a2b8; color: white; }}
                .rank-a {{ background-color: #ffc107; color: black; }}
                .rank-b {{ background-color: #6c757d; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>Research Monitoring Report</h2>
                <p>{datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            </div>
            
            <div class="stats">
                <h3>Summary</h3>
                <ul>
                    <li><strong>Query:</strong> {filters.get('query', 'N/A')}</li>
                    <li><strong>New Papers Found:</strong> {stats.get('new_papers', 0)}</li>
                    <li><strong>Papers After Filters:</strong> {len(papers)}</li>
                </ul>
                
                <h4>Venue Rank Distribution:</h4>
                <ul>
        """
        
        for rank in ["A*", "Q1", "A", "B", "Unranked"]:
            if rank in rank_counts:
                html_body += f"<li><strong>{rank}:</strong> {rank_counts[rank]} papers</li>"
        
        html_body += """
                </ul>
            </div>
            
            <h3>Top Papers:</h3>
        """
        
        # Show top 10 papers
        for i, paper in enumerate(papers[:10], 1):
            rank = paper["venue_rank"]
            rank_class = {
                "A*": "rank-as",
                "Q1": "rank-q1",
                "A": "rank-a",
                "B": "rank-b",
                "Preprint": "rank-b"
            }.get(rank, "")
            
            authors = ", ".join(paper.get("authors", [])[:3])
            if len(paper.get("authors", [])) > 3:
                authors += " et al."
            
            # Format venue with arXiv link
            venue_display = paper.get('venue', 'N/A')
            if paper.get("source") == "arxiv":
                arxiv_link = f"https://arxiv.org/abs/{paper['arxiv_id']}"
                venue_display = f'<a href="{arxiv_link}" style="color: #4472C4;">arXiv:{paper["arxiv_id"]}</a>'
            
            html_body += f"""
            <div class="paper">
                <h4>{i}. {paper['title']}</h4>
                <p>
                    <strong>Authors:</strong> {authors}<br>
                    <strong>Venue:</strong> {venue_display} 
                    <span class="badge {rank_class}">{rank}</span><br>
                    <strong>Year:</strong> {paper.get('year', 'N/A')} | 
                    <strong>Citations:</strong> {paper['citations']}
                </p>
            </div>
            """
        
        if len(papers) > 10:
            html_body += f"<p><em>...and {len(papers) - 10} more papers in the attached Excel report.</em></p>"
        
        # Active filters
        if any([filters.get("author_names"), filters.get("author_institutions"), 
                filters.get("venues"), filters.get("min_venue_rank")]):
            html_body += """
            <div class="stats">
                <h4>Active Filters:</h4>
                <ul>
            """
            
            if filters.get("author_names"):
                html_body += f"<li><strong>Authors:</strong> {', '.join(filters['author_names'])}</li>"
            if filters.get("author_institutions"):
                html_body += f"<li><strong>Institutions:</strong> {', '.join(filters['author_institutions'])}</li>"
            if filters.get("venues"):
                html_body += f"<li><strong>Venues:</strong> {', '.join(filters['venues'])}</li>"
            if filters.get("min_venue_rank"):
                html_body += f"<li><strong>Min Rank:</strong> {filters['min_venue_rank']}</li>"
            
            html_body += "</ul></div>"
        
        html_body += """
            <p style="color: #666; font-size: 12px; margin-top: 30px;">
                This is an automated notification from the Research Monitoring System.
            </p>
        </body>
        </html>
        """
        
        subject = f"{len(papers)} New Research Papers - {datetime.now().strftime('%Y-%m-%d')}"

        html_name = f"research_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        _telegram_send_document(
            html_body.encode("utf-8"),
            html_name,
            caption=subject,
        )
        if report_path and Path(report_path).exists():
            with open(report_path, "rb") as f:
                xlsx_raw = f.read()
            _telegram_send_document(
                xlsx_raw,
                Path(report_path).name,
                caption="Excel report",
            )

        logger.info("Telegram notification sent")
        return {"email_sent": True}
        
    except Exception as e:
        logger.error(f"Failed to send Telegram notification: {e}")
        return {"email_sent": False, "error": str(e)}


# ================================
# CONDITIONAL ROUTING
# ================================
def should_continue_after_fetch(state: ResearchState) -> str:
    if state.get("error"):
        return "error"
    if not state.get("papers"):
        return "end"
    return "continue"


def has_new_papers(state: ResearchState) -> str:
    """Route after dedupe. Empty fetch → skip; otherwise process (dedupe node may fill fallback batch)."""
    if state.get("new_papers"):
        return "process"
    logger.info("No papers to process (nothing fetched)")
    return "skip"


def has_filtered_papers(state: ResearchState) -> str:
    if state.get("filtered_papers"):
        return "continue"
    logger.info("No papers passed filters")
    return "end"


# ================================
# GRAPH CONSTRUCTION
# ================================
def build_workflow():
    """
    Multi-Agent Workflow with arXiv + Semantic Scholar
    
    START
      ↓
    Agent 1a: Fetch from arXiv
      ↓
    Agent 1b: Fetch from Semantic Scholar
      ↓
    Agent 2: Merge + Filter Duplicates → [none] → END
      ↓
    Agent 3: Enrich Authors (if needed)
      ↓
    Agent 4: Apply User Filters → [none passed] → END
      ↓
    Agent 5: Analyze Impact
      ↓
    Agent 6: Store Data
      ↓
    Agent 7: Generate Excel
      ↓
    Agent 8: Send Email
      ↓
    END
    """
    workflow = StateGraph(ResearchState)
    
    # Add nodes
    workflow.add_node("fetch_arxiv", fetch_arxiv_papers_node)
    workflow.add_node("fetch_semantic", fetch_papers_node)
    workflow.add_node("filter_duplicates", filter_duplicates_node)
    workflow.add_node("enrich_authors", enrich_authors_node)
    workflow.add_node("apply_filters", apply_filters_node)
    workflow.add_node("analyze_impact", analyze_impact_node)
    workflow.add_node("store", store_data_node)
    workflow.add_node("excel", generate_excel_node)
    workflow.add_node("email", send_email_node)
    
    # Define flow
    workflow.set_entry_point("fetch_arxiv")
    workflow.add_edge("fetch_arxiv", "fetch_semantic")
    
    workflow.add_conditional_edges(
        "fetch_semantic",
        should_continue_after_fetch,
        {"continue": "filter_duplicates", "end": END, "error": END}
    )
    
    workflow.add_conditional_edges(
        "filter_duplicates",
        has_new_papers,
        {"process": "enrich_authors", "skip": END}
    )
    
    workflow.add_edge("enrich_authors", "apply_filters")
    
    workflow.add_conditional_edges(
        "apply_filters",
        has_filtered_papers,
        {"continue": "analyze_impact", "end": END}
    )
    
    workflow.add_edge("analyze_impact", "store")
    workflow.add_edge("store", "excel")
    workflow.add_edge("excel", "email")
    workflow.add_edge("email", END)
    
    return workflow.compile()


# ================================
# MAIN EXECUTION
# ================================
class ResearchMonitor:
    def __init__(self):
        self.graph = build_workflow()
    
    def run(self):
        logger.info("=" * 70)
        logger.info("Starting Multi-Agent Research Monitoring Pipeline")
        logger.info("=" * 70)
        
        # Load filters
        filters = UserFilters.load()
        
        initial_state = {
            "filters": filters,
            "papers": [],
            "semantic_scholar_papers": [],
            "arxiv_papers": [],
            "new_papers": [],
            "dedupe_new_count": 0,
            "filtered_papers": [],
            "report_path": None,
            "email_sent": False,
            "error": None,
            "stats": {},
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        final_state = self.graph.invoke(initial_state)
        
        logger.info("=" * 70)
        logger.info("Pipeline Complete")
        logger.info(f"Papers found: {final_state.get('stats', {}).get('after_filters', 0)}")
        logger.info(f"Report: {final_state.get('report_path', 'N/A')}")
        logger.info(f"Email sent: {final_state.get('email_sent', False)}")
        logger.info("=" * 70)


def configure_filters_interactive():
    """Interactive CLI to configure filters"""
    print("\n" + "=" * 70)
    print("CONFIGURE RESEARCH FILTERS")
    print("=" * 70)
    
    filters = UserFilters.load()
    
    print("\nCurrent filters:")
    print(json.dumps(filters, indent=2))
    
    print("\n" + "-" * 70)
    print("Configure new filters (press Enter to skip)")
    print("-" * 70)
    
    # Query
    query = input(f"\nSearch query [{filters.get('query')}]: ").strip()
    if query:
        filters["query"] = query
    
    # Author names
    authors = input("\nAuthor names (comma-separated): ").strip()
    if authors:
        filters["author_names"] = [a.strip() for a in authors.split(",")]
    
    # Institutions
    institutions = input("\nInstitutions (comma-separated, e.g., MIT, Stanford, Google): ").strip()
    if institutions:
        filters["author_institutions"] = [i.strip() for i in institutions.split(",")]
    
    # Venues
    venues = input("\nVenues (comma-separated, e.g., CVPR, ICCV, NeurIPS): ").strip()
    if venues:
        filters["venues"] = [v.strip() for v in venues.split(",")]
    
    # Min venue rank
    print("\nMin venue rank options: A*, Q1, A, B")
    min_rank = input("Min venue rank (leave empty for any): ").strip()
    if min_rank:
        filters["min_venue_rank"] = min_rank
    
    # Citations
    min_cit = input("\nMin citations (0 for any): ").strip()
    if min_cit.isdigit():
        filters["min_citations"] = int(min_cit)
    
    # Years
    min_year = input("\nMin year (e.g., 2020): ").strip()
    if min_year.isdigit():
        filters["min_year"] = int(min_year)
    
    max_year = input("Max year (e.g., 2024): ").strip()
    if max_year.isdigit():
        filters["max_year"] = int(max_year)
    
    # Last N days
    last_days = input("\nPapers from last N days (leave empty to disable): ").strip()
    if last_days.isdigit():
        filters["last_n_days"] = int(last_days)
    
    # Keywords include
    keywords_inc = input("\nKeywords that MUST appear (comma-separated): ").strip()
    if keywords_inc:
        filters["keywords_include"] = [k.strip() for k in keywords_inc.split(",")]
    
    # Keywords exclude
    keywords_exc = input("\nKeywords to EXCLUDE (comma-separated): ").strip()
    if keywords_exc:
        filters["keywords_exclude"] = [k.strip() for k in keywords_exc.split(",")]
    
    # Max results
    max_res = input(f"\nMax results [{filters.get('max_results', 20)}]: ").strip()
    if max_res.isdigit():
        filters["max_results"] = int(max_res)
    
    # Save
    UserFilters.save(filters)
    
    print("\nFilters saved!")
    print("\nNew configuration:")
    print(json.dumps(filters, indent=2))


def main():
    import sys
    
    # Check for configuration command
    if len(sys.argv) > 1 and sys.argv[1] == "--configure":
        configure_filters_interactive()
        return
    
    if len(sys.argv) > 1 and sys.argv[1] == "--run-once":
        monitor = ResearchMonitor()
        monitor.run()
        return
    
    # Normal scheduled operation
    monitor = ResearchMonitor()
    
    # Run immediately on startup
    logger.info("Running initial execution...")
    monitor.run()
    
    # Schedule daily runs
    scheduler = BlockingScheduler()
    scheduler.add_job(
        monitor.run, 
        'cron', 
        hour=Config.RUN_HOUR, 
        minute=Config.RUN_MINUTE,
        name='daily_research_monitor'
    )
    
    logger.info(f"Scheduler active - runs daily at {Config.RUN_HOUR:02d}:{Config.RUN_MINUTE:02d}")
    logger.info("Commands:")
    logger.info("   python script.py --configure   (set filters)")
    logger.info("   python script.py --run-once    (single run)")
    logger.info("   Press Ctrl+C to stop")
    
    try:
        scheduler.start()
    except KeyboardInterrupt:
        logger.info("\nShutdown")


if __name__ == "__main__":
    main()